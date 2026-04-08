from openpyxl import Workbook

wb = Workbook()

turmas = ["2A","2B","2C","2D","3A","3B","3C","3D"]

# LISTA DE ALUNOS POR TURMA (edite conforme necessário)
alunos_por_turma = {
    "2A": ["Ana Silva", "Bruno Souza", "Carlos Lima"],
    "2B": ["Daniela Rocha", "Eduardo Alves"],
    "2C": [],
    "2D": [],
    "3A": ["Fernanda Costa", "Gabriel Martins"],
    "3B": [],
    "3C": [],
    "3D": []
}

for i, turma in enumerate(turmas):
    if i == 0:
        ws = wb.active
        ws.title = turma
    else:
        ws = wb.create_sheet(title=turma)
    
    headers = [
        "Nome",
        "N1_B1","N2_B1","N3_B1","Atv1_B1","Atv2_B1","Media_B1",
        "N1_B2","N2_B2","N3_B2","Atv1_B2","Atv2_B2","Media_B2",
        "N1_B3","N2_B3","N3_B3","Atv1_B3","Atv2_B3","Media_B3",
        "N1_B4","N2_B4","N3_B4","Atv1_B4","Atv2_B4","Media_B4",
        "Media_Anual","Status"
    ]
    
    ws.append(headers)
    
    # pega a lista de alunos da turma atual
    lista_nomes = alunos_por_turma.get(turma, [])
    
    # percorre apenas a quantidade real de alunos
    for row, nome in enumerate(lista_nomes, start=2):
        ws.cell(row=row, column=1, value=nome)
        
        medias = []

        for b in range(4):
            col_base = 2 + b*6
            
            n1 = ws.cell(row=row, column=col_base).coordinate
            n2 = ws.cell(row=row, column=col_base+1).coordinate
            n3 = ws.cell(row=row, column=col_base+2).coordinate
            atv1 = ws.cell(row=row, column=col_base+3).coordinate
            atv2 = ws.cell(row=row, column=col_base+4).coordinate
            
            media_cell = ws.cell(row=row, column=col_base+5)
            
            # média do bimestre com arredondamento para 0.5 e limite máximo 10
            media_cell.value = (
                f"=MIN(ROUND(((({n1}+{n2}+{n3})/3)+{atv1}+{atv2})*2,0)/2,10)"
            )
            
            medias.append(media_cell.coordinate)
        
        m1, m2, m3, m4 = medias
        
        media_anual = ws.cell(row=row, column=26)
        
        # média progressiva + arredondamento para 0.5
        media_anual.value = (
            f'=ROUND(('
            f'IF({m2}=0,{m1},'
            f'IF({m3}=0,AVERAGE({m1},{m2}),'
            f'IF({m4}=0,AVERAGE({m1},{m2},{m3}),'
            f'AVERAGE({m1},{m2},{m3},{m4}))))'
            f')*2,0)/2'
        )
        
        # status final
        status = ws.cell(row=row, column=27)
        status.value = f'=IF({media_anual.coordinate}>=6,"Aprovado","Exame")'

# salvar
file_path = "planilha_turmas.xlsx"
wb.save(file_path)