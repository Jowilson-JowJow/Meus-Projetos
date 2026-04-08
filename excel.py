import os
from openpyxl import Workbook

# 1. Configuração do Workbook e Turmas
wb = Workbook()
turmas = ["2A", "2B", "2C", "2D", "3A", "3B", "3C", "3D"]

for i, turma in enumerate(turmas):
    if i == 0:
        ws = wb.active
        ws.title = turma
    else:
        ws = wb.create_sheet(title=turma)
    
    # Cabeçalhos
    headers = [
        "Nome",
        "N1_B1","N2_B1","N3_B1","Atv1_B1","Atv2_B1","Media_B1",
        "N1_B2","N2_B2","N3_B2","Atv1_B2","Atv2_B2","Media_B2",
        "N1_B3","N2_B3","N3_B3","Atv1_B3","Atv2_B3","Media_B3",
        "N1_B4","N2_B4","N3_B4","Atv1_B4","Atv2_B4","Media_B4",
        "Media_Anual","Status"
    ]
    
    ws.append(headers)
    
    # 2. Preenchimento de Alunos e Fórmulas
    for row in range(2, 42):
        ws.cell(row=row, column=1, value=f"Aluno {row-1}")
        
        medias = []

        for b in range(4):
            col_base = 2 + b*6
            
            # Coleta as coordenadas das células para a fórmula
            n1 = ws.cell(row=row, column=col_base).coordinate
            n2 = ws.cell(row=row, column=col_base+1).coordinate
            n3 = ws.cell(row=row, column=col_base+2).coordinate
            atv1 = ws.cell(row=row, column=col_base+3).coordinate
            atv2 = ws.cell(row=row, column=col_base+4).coordinate
            
            media_cell = ws.cell(row=row, column=col_base+5)
            
<<<<<<< HEAD
            # média do bimestre com arredondamento para 0.5 e limite máximo 10
            media_cell.value = (
                f"=MIN(ROUND(((({n1}+{n2}+{n3})/3)+{atv1}+{atv2})*2,0)/2,10)"
            )
            
=======
            # Média do bimestre (Fórmula Excel padrão Inglês)
            # Nota: Dividir por 3 e somar atividades
            media_cell.value = f"=(({n1}+{n2}+{n3})/3)+{atv1}+{atv2}"
>>>>>>> fcda3111134d135b69850d1239dc2be7af45a34c
            medias.append(media_cell.coordinate)
        
        m1, m2, m3, m4 = medias
        media_anual = ws.cell(row=row, column=26)
        
<<<<<<< HEAD
        # média progressiva + arredondamento para 0.5 (mantida)
=======
        # Lógica progressiva para Média Anual (Trata zeros como bimestre não cursado)
>>>>>>> fcda3111134d135b69850d1239dc2be7af45a34c
        media_anual.value = (
            f'=ROUND(('
            f'IF({m2}=0,{m1},'
            f'IF({m3}=0,AVERAGE({m1},{m2}),'
            f'IF({m4}=0,AVERAGE({m1},{m2},{m3}),'
            f'AVERAGE({m1},{m2},{m3},{m4}))))'
            f')*2,0)/2'
        )
        
        # Status final
        status = ws.cell(row=row, column=27)
        status.value = f'=IF({media_anual.coordinate}>=6,"Aprovado","Exame")'

# 3. Lógica de Salvamento Robusta
nome_arquivo = "planilha_turmas.xlsx"

# Obtém o caminho absoluto da pasta onde o script está
diretorio_script = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(diretorio_script, nome_arquivo)

# CORREÇÃO: Garante que a pasta de destino realmente existe
if not os.path.exists(diretorio_script):
    os.makedirs(diretorio_script)

try:
    # Tenta salvar o arquivo
    wb.save(file_path)
    print("-" * 50)
    print(f"SUCESSO: Arquivo criado em:\n{file_path}")
    print("-" * 50)
except PermissionError:
    print("ERRO: O arquivo já está aberto no Excel. Feche-o e tente novamente.")
except Exception as e:
    print(f"ERRO AO SALVAR: {e}")
    